from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from ..database import get_db
from ..models import Student, Class, Admin, Score
from ..schemas import StudentCreate, StudentOut, StudentBatchUpdate
from ..auth import get_current_admin, require_school, hash_password
import openpyxl
from io import BytesIO
from typing import Optional

router = APIRouter(prefix="/api/students", tags=["students"])

def _school_students(db: Session, school_id: int):
    return db.query(Student).join(Class).filter(Class.school_id == school_id)

@router.get("", response_model=list[StudentOut])
def list_students(
    search: Optional[str] = None,
    class_id: Optional[int] = None,
    page: int = 1,
    page_size: int = 50,
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin)
):
    sid = require_school(current)
    if sid is not None:
        q = _school_students(db, sid)
    else:
        q = db.query(Student)
    if search:
        q = q.filter(
            (Student.student_id.contains(search)) | (Student.name.contains(search))
        )
    if class_id:
        q = q.filter(Student.class_id == class_id)
    students = q.order_by(Student.student_id).offset((page - 1) * page_size).limit(page_size).all()

    result = []
    for s in students:
        out = StudentOut.model_validate(s)
        if s.class_:
            out.class_name = s.class_.name
            out.class_grade = s.class_.grade
        result.append(out)
    return result

@router.post("", response_model=StudentOut)
def create_student(data: StudentCreate, db: Session = Depends(get_db), current: Admin = Depends(get_current_admin)):
    sid = require_school(current)
    q = db.query(Class).filter(Class.id == data.class_id)
    if sid is not None:
        q = q.filter(Class.school_id == sid)
    cls = q.first()
    if not cls:
        raise HTTPException(status_code=400, detail="班级不存在或不属于当前学校")
    existing = db.query(Student).filter(Student.student_id == data.student_id).first()
    if existing:
        raise HTTPException(status_code=400, detail="学号已存在")
    student = Student(
        student_id=data.student_id,
        name=data.name,
        gender=data.gender,
        class_id=data.class_id,
        password_hash=hash_password(data.student_id[-6:])
    )
    db.add(student)
    db.commit()
    db.refresh(student)
    return StudentOut.model_validate(student)

@router.get("/{student_id}", response_model=StudentOut)
def get_student(student_id: int, db: Session = Depends(get_db), current: Admin = Depends(get_current_admin)):
    sid = require_school(current)
    if sid is not None:
        s = _school_students(db, sid).filter(Student.id == student_id).first()
    else:
        s = db.query(Student).filter(Student.id == student_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="学生不存在")
    out = StudentOut.model_validate(s)
    if s.class_:
        out.class_name = s.class_.name
        out.class_grade = s.class_.grade
    return out

@router.put("/{student_id}", response_model=StudentOut)
def update_student(student_id: int, data: StudentCreate, db: Session = Depends(get_db), current: Admin = Depends(get_current_admin)):
    sid = require_school(current)
    if sid is not None:
        s = _school_students(db, sid).filter(Student.id == student_id).first()
    else:
        s = db.query(Student).filter(Student.id == student_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="学生不存在")
    s.student_id = data.student_id
    s.name = data.name
    s.gender = data.gender
    s.class_id = data.class_id
    db.commit()
    db.refresh(s)
    return StudentOut.model_validate(s)

@router.delete("/{student_id}")
def delete_student(student_id: int, db: Session = Depends(get_db), current: Admin = Depends(get_current_admin)):
    sid = require_school(current)
    if sid is not None:
        s = _school_students(db, sid).filter(Student.id == student_id).first()
    else:
        s = db.query(Student).filter(Student.id == student_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="学生不存在")
    db.query(Score).filter(Score.student_id == student_id).delete()
    db.delete(s)
    db.commit()
    return {"ok": True}

@router.post("/batch-import")
def batch_import(file: UploadFile = File(...), db: Session = Depends(get_db), current: Admin = Depends(get_current_admin)):
    sid = require_school(current)
    contents = file.file.read()
    wb = openpyxl.load_workbook(BytesIO(contents))
    ws = wb.active
    imported = 0
    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0]:
            continue
        student_id = str(row[0]).strip()
        name = str(row[1]).strip()
        gender_str = str(row[2]).strip()
        class_str = str(row[3]).strip()
        gender = "M" if gender_str == "男" else "F"
        if "届" in class_str:
            parts = class_str.split("届", 1)
            grade = parts[0] + "届"
            class_name = parts[1]
        else:
            grade = class_str
            class_name = ""
        q = db.query(Class).filter(Class.grade == grade, Class.name == class_name)
        if sid is not None:
            q = q.filter(Class.school_id == sid)
        cls = q.first()
        if not cls:
            if sid is None:
                errors.append(f"行{row_idx}: 超管导入需指定现有班级（无法自动创建跨学校班级）")
                continue
            cls = Class(grade=grade, name=class_name, school_id=sid)
            db.add(cls)
            db.flush()
        existing = db.query(Student).filter(Student.student_id == student_id).first()
        if existing:
            errors.append(f"行{row_idx}: 学号{student_id}已存在，跳过")
            continue
        student = Student(
            student_id=student_id,
            name=name,
            gender=gender,
            class_id=cls.id,
            password_hash=hash_password(student_id[-6:])
        )
        db.add(student)
        imported += 1
    db.commit()
    return {"imported": imported, "errors": errors}

@router.put("/batch/update")
def batch_update(data: StudentBatchUpdate, db: Session = Depends(get_db), current: Admin = Depends(get_current_admin)):
    sid = require_school(current)
    if sid is not None:
        q = _school_students(db, sid)
    else:
        q = db.query(Student)
    if data.class_id:
        q = q.filter(Student.class_id == data.class_id)
    students = q.all()
    count = 0
    for s in students:
        if data.new_class_id:
            s.class_id = data.new_class_id
        if data.reset_password:
            s.password_hash = hash_password(s.student_id[-6:])
        count += 1
    db.commit()
    return {"updated": count}

@router.delete("/batch-delete")
def batch_delete_students(
    ids: list[int] = [],
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin)
):
    """Batch delete students and their scores."""
    if not ids:
        raise HTTPException(status_code=400, detail="请选择要删除的学生")
    db.query(Score).filter(Score.student_id.in_(ids)).delete(synchronize_session=False)
    db.query(Student).filter(Student.id.in_(ids)).delete(synchronize_session=False)
    db.commit()
    return {"ok": True, "deleted": len(ids)}

@router.get("/template/download")
def download_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "学生信息"
    ws.append(["学号", "姓名", "性别", "班级"])
    ws.append(["270301", "张三", "女", "2027届3班"])
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=student_template.xlsx"}
    )
