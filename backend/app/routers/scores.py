from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from ..database import get_db
from ..models import Score, Student, SportEvent, ScoringStandard, Class, Admin, SystemConfig, InputFormat, Gender
from ..schemas import ScoreBatchSave, ScoreWithChange, ClearAllRequest
from ..auth import get_current_admin, get_current_admin_flexible, get_super_admin, get_school_admin, require_school, verify_password
from ..scoring import calculate_score, normalize_time_ms, parse_value
from ..config import settings
import openpyxl
from io import BytesIO
from datetime import date, timedelta
from typing import Optional
from collections import defaultdict

router = APIRouter(prefix="/api/scores", tags=["scores"])

def _get_admin_school(current: Admin) -> Optional[int]:
    return require_school(current)

def _maybe_filter(query, model, sid):
    """Apply school filter only if sid is not None (super-admin sees all)."""
    if sid is not None:
        return query.filter(model.school_id == sid)
    return query

def _pick_best_in_window(scores, days=30):
    """Keep best earned_score per (student_id, event_id) within the last `days` days.
    Returns dict {(student_id, event_id): Score}"""
    cutoff = date.today() - timedelta(days=days)
    best = {}
    for sc in scores:
        if sc.test_date < cutoff:
            continue
        key = (sc.student_id, sc.event_id)
        if key not in best or sc.earned_score > best[key].earned_score:
            best[key] = sc
    return best

def _calc_top3_scores(student_event_scores: dict, mandatory_event_ids: set) -> list:
    """中考总分 = 长跑(800/1000米必考)最好成绩 + 其余项目最好2项之和（满分30）。

    student_event_scores: {student_id: {event_id: score}}
    mandatory_event_ids: 必考项目 event_id 集合
    """
    result = []
    for evt_scores in student_event_scores.values():
        mandatory = []
        other = []
        for eid, score in evt_scores.items():
            if eid in mandatory_event_ids:
                mandatory.append(score)
            else:
                other.append(score)
        best_mandatory = max(mandatory) if mandatory else 0
        best2 = sum(sorted(other, reverse=True)[:2])
        result.append(best_mandatory + best2)
    return result

def _score_distribution(totals: list, participants: int) -> dict:
    """将中考总分分档统计，返回满分率和各档位人数"""
    buckets = {"30(满分)": 0, "28-29": 0, "25-27": 0, "20-24": 0, "20以下": 0}
    for t in totals:
        if t >= 30:
            buckets["30(满分)"] += 1
        elif t >= 28:
            buckets["28-29"] += 1
        elif t >= 25:
            buckets["25-27"] += 1
        elif t >= 20:
            buckets["20-24"] += 1
        else:
            buckets["20以下"] += 1
    return {
        "full_score_rate": round(buckets["30(满分)"] / participants * 100, 1) if participants else 0,
        "buckets": [{"label": k, "count": v} for k, v in buckets.items()]
    }

def _get_previous_score(db: Session, student_db_id: int, event_id: int, current_date: date) -> Optional[Score]:
    return (
        db.query(Score)
        .filter(
            Score.student_id == student_db_id,
            Score.event_id == event_id,
            Score.test_date < current_date
        )
        .order_by(Score.test_date.desc())
        .first()
    )

@router.post("/batch", response_model=list[ScoreWithChange])
def batch_save_scores(
    data: ScoreBatchSave,
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin)
):
    sid = _get_admin_school(current)
    if sid is None:
        raise HTTPException(status_code=400, detail="请先选择学校再进行成绩录入")
    results = []
    praise_threshold = 1
    warning_threshold = 2
    praise_cfg = _maybe_filter(db.query(SystemConfig), SystemConfig, sid).filter(
        SystemConfig.key == "praise_threshold"
    ).first()
    warning_cfg = _maybe_filter(db.query(SystemConfig), SystemConfig, sid).filter(
        SystemConfig.key == "warning_threshold"
    ).first()
    if praise_cfg:
        praise_threshold = int(praise_cfg.value)
    if warning_cfg:
        warning_threshold = int(warning_cfg.value)

    for entry in data.scores:
        event = _maybe_filter(
            db.query(SportEvent).filter(SportEvent.id == entry.event_id),
            SportEvent, sid
        ).first()
        student = db.query(Student).get(entry.student_id)
        raw_value = entry.raw_value
        if event and event.input_format == InputFormat.time_ms:
            raw_value = normalize_time_ms(raw_value)
        standards = db.query(ScoringStandard).filter(ScoringStandard.event_id == entry.event_id).all()
        earned = calculate_score(raw_value, event, standards, student.gender.value if student else None)

        prev = _get_previous_score(db, entry.student_id, entry.event_id, entry.test_date)
        prev_score = None
        change = None
        is_praise = False
        is_warning = False
        if prev:
            prev_score = prev.earned_score
            change = earned - prev_score
            is_praise = change >= praise_threshold
            is_warning = (prev_score - earned) >= warning_threshold

        existing = (
            db.query(Score)
            .filter(
                Score.student_id == entry.student_id,
                Score.event_id == entry.event_id,
                Score.test_date == entry.test_date
            ).first()
        )
        if existing:
            existing.raw_value = raw_value
            existing.earned_score = earned
            existing.recorder_id = current.id
            score_obj = existing
        else:
            score_obj = Score(
                student_id=entry.student_id,
                event_id=entry.event_id,
                raw_value=raw_value,
                earned_score=earned,
                test_date=entry.test_date,
                recorder_id=current.id,
                school_id=sid
            )
            db.add(score_obj)

        db.flush()
        db.refresh(score_obj)

        result = ScoreWithChange(
            id=score_obj.id,
            student_id=score_obj.student_id,
            event_id=score_obj.event_id,
            raw_value=score_obj.raw_value,
            earned_score=score_obj.earned_score,
            test_date=score_obj.test_date,
            previous_score=prev_score,
            change=change,
            is_praise=is_praise,
            is_warning=is_warning,
        )
        results.append(result)

    db.commit()
    return results

@router.get("/batch-import-template")
def download_batch_import_template(
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin)
):
    """Download an Excel template for batch score import with event columns."""
    sid = _get_admin_school(current)
    if sid is None:
        raise HTTPException(400, "请先选择学校再进行操作")

    events = db.query(SportEvent).filter(SportEvent.school_id == sid).order_by(SportEvent.sort_order).all()
    if not events:
        raise HTTPException(400, "当前学校没有体育项目，请先在项目设置中添加")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "成绩导入模板"

    headers = ["学号", "姓名"] + [e.name for e in events]
    ws.append(headers)

    # Example row
    example = ["202401", "张三"] + [""] * len(events)
    ws.append(example)

    # Format: auto-width for readability
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = openpyxl.styles.Font(bold=True)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=score_import_template.xlsx"}
    )


@router.post("/batch-import")
def batch_import_scores(
    class_id: int = Query(...),
    test_date: str = Query(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin)
):
    """Import scores from Excel. Col A=student_id, Col B=name, Col C+=event names as headers."""
    sid = _get_admin_school(current)

    cls = db.query(Class).filter(Class.id == class_id, Class.school_id == sid).first()
    if not cls:
        raise HTTPException(404, "班级不存在")

    test_dt = date.fromisoformat(test_date)

    try:
        wb = openpyxl.load_workbook(file.file, data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(400, "无法读取Excel文件")

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "Excel文件为空或无数据行")

    headers = [str(h).strip() if h else "" for h in rows[0]]
    if len(headers) < 2:
        raise HTTPException(400, "表头至少需要学号和姓名两列")

    # Match header columns to sport events (columns C+)
    events = db.query(SportEvent).filter(SportEvent.school_id == sid).all()
    col_event_map = {}  # col_idx -> SportEvent
    for col_idx, h in enumerate(headers):
        if col_idx < 2:
            continue  # skip 学号, 姓名
        for e in events:
            if h == e.name:
                col_event_map[col_idx] = e
                break

    if not col_event_map:
        raise HTTPException(400, "未找到匹配的体育项目，请检查表头是否包含项目名称")

    # Preload class students for matching
    students = db.query(Student).filter(Student.class_id == class_id).all()
    student_map = {s.student_id: s for s in students}

    imported = 0
    skipped = []
    for row_idx, row in enumerate(rows[1:], start=2):
        student_id = str(row[0]).strip() if row[0] else ""
        student = student_map.get(student_id)
        if not student:
            skipped.append(f"行{row_idx}: 学号{student_id}不在本班")
            continue

        for col_idx, event in col_event_map.items():
            raw_value = row[col_idx]
            if raw_value is None or str(raw_value).strip() == "":
                continue
            raw_value = str(raw_value).strip()

            # Normalize time formats
            if event.input_format == InputFormat.time_ms:
                raw_value = normalize_time_ms(raw_value)

            standards = db.query(ScoringStandard).filter(
                ScoringStandard.event_id == event.id
            ).all()
            earned = calculate_score(raw_value, event, standards, student.gender.value)

            # Upsert
            existing = db.query(Score).filter(
                Score.student_id == student.id,
                Score.event_id == event.id,
                Score.test_date == test_dt
            ).first()
            if existing:
                existing.raw_value = raw_value
                existing.earned_score = earned
                existing.recorder_id = current.id
            else:
                db.add(Score(
                    student_id=student.id,
                    event_id=event.id,
                    raw_value=raw_value,
                    earned_score=earned,
                    test_date=test_dt,
                    recorder_id=current.id,
                    school_id=sid
                ))
            imported += 1

    db.commit()
    return {"ok": True, "imported": imported, "skipped": skipped}

@router.delete("/{score_id}")
def delete_score(score_id: int, db: Session = Depends(get_db), current: Admin = Depends(get_current_admin)):
    sc = db.query(Score).get(score_id)
    if not sc:
        raise HTTPException(404, "成绩记录不存在")
    db.delete(sc)
    db.commit()
    return {"ok": True}

@router.get("/class-stats")
def class_stats(
    class_id: int = Query(...),
    event_ids: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin)
):
    sid = _get_admin_school(current)
    cls = _maybe_filter(db.query(Class), Class, sid).filter(Class.id == class_id).first()
    if not cls:
        raise HTTPException(404, "班级不存在")

    event_id_list = [int(x) for x in event_ids.split(",")] if event_ids else None
    students = db.query(Student).filter(Student.class_id == class_id).all()
    scores_q = _maybe_filter(db.query(Score), Score, sid).filter(
        Score.student_id.in_([s.id for s in students])
    )
    if event_id_list:
        scores_q = scores_q.filter(Score.event_id.in_(event_id_list))

    all_scores = scores_q.order_by(Score.test_date.desc()).all()
    best = _pick_best_in_window(all_scores)

    events = _maybe_filter(db.query(SportEvent), SportEvent, sid).order_by(SportEvent.sort_order).all()
    # 必考项：长跑 800/1000 米（从全校全部项目中识别，不受 event_ids 筛选影响）
    mandatory_event_ids = {e.id for e in events if '800' in e.name or '1000' in e.name}
    if event_id_list:
        events = [e for e in events if e.id in event_id_list]

    event_scores = defaultdict(list)
    student_totals = defaultdict(list)
    student_event_scores = defaultdict(dict)  # {student_id: {event_id: score}}
    for (_sid, eid), sc in best.items():
        event_scores[eid].append(sc.earned_score)
        student_totals[_sid].append(sc.earned_score)
        student_event_scores[_sid][eid] = sc.earned_score

    event_avgs = []
    for e in events:
        scores_list = event_scores.get(e.id, [])
        avg = sum(scores_list) / len(scores_list) if scores_list else 0
        event_avgs.append({"event_id": e.id, "event_name": e.name, "avg_score": round(avg, 1), "count": len(scores_list)})

    all_scores_list = [sum(v) for v in student_totals.values() if v]
    max_per_student = len(events)
    n_participants = len(student_totals)
    overall_avg = sum(all_scores_list) / len(all_scores_list) if all_scores_list else 0
    excellent_count = sum(1 for t in all_scores_list if max_per_student > 0 and t / max_per_student >= 9)
    pass_count = sum(1 for t in all_scores_list if max_per_student > 0 and t / max_per_student >= 6)

    # 中考总分 = 长跑(必考)最好成绩 + 其余项目最好2项之和
    total_scores_3best = _calc_top3_scores(student_event_scores, mandatory_event_ids)
    dist = _score_distribution(total_scores_3best, n_participants)

    warning_students = []
    for s in students:
        for e in events:
            student_scores = sorted(
                [sc for sc in all_scores if sc.student_id == s.id and sc.event_id == e.id],
                key=lambda x: x.test_date
            )
            if len(student_scores) >= 2:
                prev_score = student_scores[-2].earned_score
                curr_score = student_scores[-1].earned_score
                if prev_score - curr_score >= 2:
                    warning_students.append({
                        "student_id": s.id,
                        "student_name": s.name,
                        "student_no": s.student_id,
                        "student_gender": s.gender.value,
                        "class_name": f"{cls.grade}{cls.name}",
                        "event_name": e.name,
                        "prev_score": prev_score,
                        "curr_score": curr_score
                    })

    return {
        "class_id": class_id,
        "class_name": f"{cls.grade}{cls.name}",
        "total_students": len(students),
        "participants": n_participants,
        "avg_score": round(overall_avg, 1),
        "excellent_rate": round(excellent_count / n_participants * 100, 1) if n_participants else 0,
        "pass_rate": round(pass_count / n_participants * 100, 1) if n_participants else 0,
        "full_score_rate": dist["full_score_rate"],
        "score_distribution": dist["buckets"],
        "event_avgs": event_avgs,
        "warning_students": warning_students
    }

@router.get("/student-stats/{student_id}")
def student_stats(
    student_id: int,
    event_ids: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin)
):
    sid = _get_admin_school(current)
    s = db.query(Student).join(Class).filter(
        Student.id == student_id, Class.school_id == sid
    ).first()
    if not s:
        raise HTTPException(404, "学生不存在")

    event_id_list = [int(x) for x in event_ids.split(",")] if event_ids else None
    scores_q = db.query(Score).filter(Score.student_id == student_id)
    if event_id_list:
        scores_q = scores_q.filter(Score.event_id.in_(event_id_list))

    all_scores = scores_q.order_by(Score.test_date.desc()).all()

    scores_by_event = defaultdict(list)
    for sc in all_scores:
        event = db.query(SportEvent).get(sc.event_id)
        try:
            nv = parse_value(sc.raw_value, event.input_format)
        except (ValueError, AttributeError):
            nv = None
        scores_by_event[event.name].append({
            "id": sc.id,
            "raw_value": sc.raw_value,
            "earned_score": sc.earned_score,
            "test_date": sc.test_date.isoformat(),
            "numeric_value": nv,
            "unit": event.unit,
            "higher_better": event.higher_better
        })

    latest_per_event = {}
    for sc in all_scores:
        if sc.event_id not in latest_per_event:
            latest_per_event[sc.event_id] = sc

    recs = sorted(latest_per_event.items(), key=lambda x: x[1].earned_score, reverse=True)[:4]
    recommended = []
    medals = ["🥇", "🥈", "🥉", "④"]
    for i, (eid, sc) in enumerate(recs):
        event = db.query(SportEvent).get(eid)
        recommended.append({
            "rank": i + 1,
            "medal": medals[i],
            "event_name": event.name,
            "score": sc.earned_score
        })

    return {
        "student": {
            "id": s.id,
            "student_id": s.student_id,
            "name": s.name,
            "gender": s.gender.value,
            "class_name": s.class_.name if s.class_ else "",
            "class_grade": s.class_.grade if s.class_ else "",
        },
        "scores_by_event": dict(scores_by_event),
        "recommended_events": recommended
    }

@router.get("/export/class")
def export_class_scores(
    class_id: int = Query(...),
    event_ids: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin)
):
    sid = _get_admin_school(current)
    cls = db.query(Class).filter(Class.id == class_id, Class.school_id == sid).first()
    if not cls:
        raise HTTPException(404, "班级不存在")
    students = db.query(Student).filter(Student.class_id == class_id).order_by(Student.student_id).all()
    event_id_list = [int(x) for x in event_ids.split(",")] if event_ids else None
    events_q = db.query(SportEvent).filter(SportEvent.school_id == sid)
    if event_id_list:
        events_q = events_q.filter(SportEvent.id.in_(event_id_list))
    events = events_q.order_by(SportEvent.sort_order).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{cls.grade}{cls.name}成绩"
    headers = ["学号", "姓名", "性别"] + [e.name for e in events] + ["总分"]
    ws.append(headers)

    for s in students:
        row = [s.student_id, s.name, "男" if s.gender.value == "M" else "女"]
        total = 0
        for e in events:
            latest = (
                db.query(Score)
                .filter(Score.student_id == s.id, Score.event_id == e.id)
                .order_by(Score.test_date.desc())
                .first()
            )
            if latest:
                row.append(latest.earned_score)
                total += latest.earned_score
            else:
                row.append("-")
        row.append(total)
        ws.append(row)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=class_{class_id}_scores.xlsx"}
    )

@router.get("/export/student/{student_id}")
def export_student_scores(
    student_id: int,
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin)
):
    s = db.query(Student).get(student_id)
    if not s:
        raise HTTPException(404)

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "成绩汇总"
    ws1.append(["项目", "成绩", "得分", "测试日期"])
    events = db.query(SportEvent).order_by(SportEvent.sort_order).all()
    total = 0
    for e in events:
        latest = (
            db.query(Score)
            .filter(Score.student_id == student_id, Score.event_id == e.id)
            .order_by(Score.test_date.desc())
            .first()
        )
        if latest:
            ws1.append([e.name, latest.raw_value, latest.earned_score, latest.test_date.isoformat()])
            total += latest.earned_score
        else:
            ws1.append([e.name, "-", "-", "-"])
    ws1.append(["总分", "", total, ""])

    ws2 = wb.create_sheet("历史记录")
    ws2.append(["项目", "成绩", "得分", "测试日期"])
    scores = db.query(Score).filter(Score.student_id == student_id).order_by(Score.test_date.desc()).all()
    for sc in scores:
        event = db.query(SportEvent).get(sc.event_id)
        ws2.append([event.name, sc.raw_value, sc.earned_score, sc.test_date.isoformat()])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=student_{student_id}_scores.xlsx"}
    )

@router.get("/school-stats")
def school_stats(
    event_ids: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin)
):
    """School-wide statistics."""
    sid = _get_admin_school(current)
    event_id_list = [int(x) for x in event_ids.split(",")] if event_ids else None
    events = _maybe_filter(db.query(SportEvent), SportEvent, sid).order_by(SportEvent.sort_order).all()
    # 必考项：长跑 800/1000 米（从全校全部项目中识别）
    mandatory_event_ids = {e.id for e in events if '800' in e.name or '1000' in e.name}
    if event_id_list:
        events = [e for e in events if e.id in event_id_list]

    all_students = _maybe_filter(db.query(Student).join(Class), Class, sid).all()
    all_scores = _maybe_filter(db.query(Score), Score, sid).order_by(Score.test_date.desc()).all()

    # Best score per student per event within last 30 days
    best = _pick_best_in_window(all_scores)

    event_scores = defaultdict(list)
    student_totals = defaultdict(list)
    student_event_scores = defaultdict(dict)
    for (_sid, eid), sc in best.items():
        if eid in [e.id for e in events]:
            event_scores[eid].append(sc.earned_score)
            student_totals[_sid].append(sc.earned_score)
            student_event_scores[_sid][eid] = sc.earned_score

    event_avgs = []
    for e in events:
        scores_list = event_scores.get(e.id, [])
        avg = sum(scores_list) / len(scores_list) if scores_list else 0
        event_avgs.append({"event_id": e.id, "event_name": e.name, "avg_score": round(avg, 1), "count": len(scores_list)})

    all_scores_list = [sum(v) for v in student_totals.values() if v]
    n_participants = len(student_totals)
    max_per_student = len(events)
    overall_avg = sum(all_scores_list) / len(all_scores_list) if all_scores_list else 0
    excellent_count = sum(1 for t in all_scores_list if max_per_student > 0 and t / max_per_student >= 9)
    pass_count = sum(1 for t in all_scores_list if max_per_student > 0 and t / max_per_student >= 6)

    # 中考总分 = 长跑(必考)最好成绩 + 其余项目最好2项之和
    total_scores_3best = _calc_top3_scores(student_event_scores, mandatory_event_ids)
    dist = _score_distribution(total_scores_3best, n_participants)

    classes = _maybe_filter(db.query(Class), Class, sid).order_by(Class.grade, Class.name).all()
    # Build student-id -> class lookup for warning
    student_class_map = {}
    for s in all_students:
        student_class_map[s.id] = f"{s.class_.grade}{s.class_.name}" if s.class_ else ""

    class_summaries = []
    for cls in classes:
        cls_students = [s for s in all_students if s.class_id == cls.id]
        cls_total = 0
        cls_count = 0
        for s in cls_students:
            if s.id in student_totals:
                cls_total += sum(student_totals[s.id])
                cls_count += 1
        cls_avg = cls_total / (cls_count * max_per_student) * 10 if cls_count > 0 and max_per_student > 0 else 0
        class_summaries.append({
            "class_id": cls.id, "class_name": f"{cls.grade}{cls.name}",
            "students": len(cls_students), "participants": cls_count,
            "avg_score": round(cls_avg, 1)
        })

    warning_students = []
    for s in all_students:
        for e in events:
            student_scores = sorted(
                [sc for sc in all_scores if sc.student_id == s.id and sc.event_id == e.id],
                key=lambda x: x.test_date
            )
            if len(student_scores) >= 2:
                prev = student_scores[-2].earned_score
                curr = student_scores[-1].earned_score
                if prev - curr >= 2:
                    warning_students.append({
                        "student_no": s.student_id, "student_name": s.name,
                        "student_gender": s.gender.value,
                        "class_name": student_class_map.get(s.id, ""),
                        "event_name": e.name, "prev_score": prev, "curr_score": curr
                    })

    return {
        "total_students": len(all_students),
        "total_classes": len(class_summaries),
        "participants": n_participants,
        "avg_score": round(overall_avg, 1),
        "excellent_rate": round(excellent_count / n_participants * 100, 1) if n_participants else 0,
        "pass_rate": round(pass_count / n_participants * 100, 1) if n_participants else 0,
        "full_score_rate": dist["full_score_rate"],
        "score_distribution": dist["buckets"],
        "event_avgs": event_avgs,
        "class_summaries": class_summaries,
        "warning_students": warning_students
    }


@router.get("/grade-stats")
def grade_stats(
    event_ids: Optional[str] = Query(None),
    grade: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin)
):
    """Grade-level statistics: group classes by grade and aggregate."""
    sid = _get_admin_school(current)
    event_id_list = [int(x) for x in event_ids.split(",")] if event_ids else None
    events = _maybe_filter(db.query(SportEvent), SportEvent, sid).order_by(SportEvent.sort_order).all()
    # 必考项：长跑 800/1000 米（从全校全部项目中识别）
    mandatory_event_ids = {e.id for e in events if '800' in e.name or '1000' in e.name}
    if event_id_list:
        events = [e for e in events if e.id in event_id_list]

    all_classes = _maybe_filter(db.query(Class), Class, sid).order_by(Class.grade, Class.name).all()
    all_scores = _maybe_filter(db.query(Score), Score, sid).order_by(Score.test_date.desc()).all()

    # Best score per student per event within last 30 days
    best = _pick_best_in_window(all_scores)

    # Group classes by grade
    grade_classes = defaultdict(list)
    for cls in all_classes:
        grade_classes[cls.grade].append(cls)

    if grade:
        grade_classes = {grade: grade_classes[grade]} if grade in grade_classes else {}

    result = []
    for grade_name, classes_in_grade in sorted(grade_classes.items()):
        class_ids = [c.id for c in classes_in_grade]
        grade_students = db.query(Student).filter(Student.class_id.in_(class_ids)).all()
        grade_student_ids = [s.id for s in grade_students]

        event_scores = defaultdict(list)
        student_totals = defaultdict(list)
        student_event_scores = defaultdict(dict)
        for (st_id, eid), sc in best.items():
            if st_id in grade_student_ids and eid in [e.id for e in events]:
                event_scores[eid].append(sc.earned_score)
                student_totals[st_id].append(sc.earned_score)
                student_event_scores[st_id][eid] = sc.earned_score

        event_avgs = []
        for e in events:
            scores_list = event_scores.get(e.id, [])
            avg = sum(scores_list) / len(scores_list) if scores_list else 0
            event_avgs.append({"event_id": e.id, "event_name": e.name, "avg_score": round(avg, 1), "count": len(scores_list)})

        all_scores_for_grade = [sum(v) for v in student_totals.values() if v]
        n_participants = len(student_totals)
        max_per_student = len(events)
        overall_avg = sum(all_scores_for_grade) / len(all_scores_for_grade) if all_scores_for_grade else 0
        excellent_count = sum(1 for t in all_scores_for_grade if max_per_student > 0 and t / max_per_student >= 9)
        pass_count = sum(1 for t in all_scores_for_grade if max_per_student > 0 and t / max_per_student >= 6)

        # 中考总分 = 长跑(必考)最好成绩 + 其余项目最好2项之和
        total_scores_3best = _calc_top3_scores(student_event_scores, mandatory_event_ids)
        dist = _score_distribution(total_scores_3best, n_participants)

        class_summaries = []
        for cls in classes_in_grade:
            cls_students = [s for s in grade_students if s.class_id == cls.id]
            cls_total = 0
            cls_count = 0
            for s in cls_students:
                if s.id in student_totals:
                    cls_total += sum(student_totals[s.id])
                    cls_count += 1
            cls_avg = cls_total / (cls_count * max_per_student) * 10 if cls_count > 0 and max_per_student > 0 else 0
            class_summaries.append({
                "class_id": cls.id, "class_name": f"{cls.grade}{cls.name}",
                "students": len(cls_students), "participants": cls_count,
                "avg_score": round(cls_avg, 1)
            })

        warning_students = []
        for s in grade_students:
            for e in events:
                student_scores = sorted(
                    [sc for sc in all_scores if sc.student_id == s.id and sc.event_id == e.id],
                    key=lambda x: x.test_date
                )
                if len(student_scores) >= 2:
                    prev = student_scores[-2].earned_score
                    curr = student_scores[-1].earned_score
                    if prev - curr >= 2:
                        cls_name = ""
                        for c in classes_in_grade:
                            if c.id == s.class_id:
                                cls_name = f"{c.grade}{c.name}"
                                break
                        warning_students.append({
                            "student_no": s.student_id, "student_name": s.name,
                            "student_gender": s.gender.value,
                            "class_name": cls_name,
                            "event_name": e.name, "prev_score": prev, "curr_score": curr
                        })

        result.append({
            "grade": grade_name,
            "total_students": len(grade_students),
            "total_classes": len(classes_in_grade),
            "participants": n_participants,
            "avg_score": round(overall_avg, 1),
            "excellent_rate": round(excellent_count / n_participants * 100, 1) if n_participants else 0,
            "pass_rate": round(pass_count / n_participants * 100, 1) if n_participants else 0,
            "full_score_rate": dist["full_score_rate"],
            "score_distribution": dist["buckets"],
            "event_avgs": event_avgs,
            "class_summaries": class_summaries,
            "warning_students": warning_students
        })

    return result


@router.post("/export/preview")
def export_preview(
    scope: str = Query(...),        # "school" | "class" | "student"
    class_id: Optional[int] = Query(None),
    student_id: Optional[int] = Query(None),
    event_ids: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    mode: str = Query("all"),       # "all" | "best" | "latest"
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin)
):
    """Preview export data before downloading."""
    sid = _get_admin_school(current)
    event_id_list = [int(x) for x in event_ids.split(",")] if event_ids else None

    # Build query - always scoped to school
    q = _maybe_filter(db.query(Score), Score, sid)
    if scope == "class" and class_id:
        students_in_class = db.query(Student).filter(Student.class_id == class_id).all()
        q = q.filter(Score.student_id.in_([s.id for s in students_in_class]))
    elif scope == "student" and student_id:
        q = q.filter(Score.student_id == student_id)
    if event_id_list:
        q = q.filter(Score.event_id.in_(event_id_list))
    if date_from:
        q = q.filter(Score.test_date >= date.fromisoformat(date_from))
    if date_to:
        q = q.filter(Score.test_date <= date.fromisoformat(date_to))

    all_scores = q.order_by(Score.test_date.desc(), Score.student_id).all()
    events = db.query(SportEvent).filter(SportEvent.school_id == sid).all()
    event_map = {e.id: e for e in events}

    if mode == "best":
        best = {}
        for sc in all_scores:
            key = (sc.student_id, sc.event_id)
            if key not in best or sc.earned_score > best[key].earned_score:
                best[key] = sc
        all_scores = sorted(best.values(), key=lambda x: (x.student_id, x.event_id))
    elif mode == "latest":
        latest = {}
        for sc in all_scores:
            key = (sc.student_id, sc.event_id)
            if key not in latest:
                latest[key] = sc
        all_scores = sorted(latest.values(), key=lambda x: (x.student_id, x.event_id))

    # Preload students and classes
    student_ids = {sc.student_id for sc in all_scores} if all_scores else set()
    if student_ids:
        students_list = db.query(Student).filter(Student.id.in_(student_ids)).all()
        class_ids = {s.class_id for s in students_list}
        classes_map = {c.id: c for c in db.query(Class).filter(Class.id.in_(class_ids)).all()} if class_ids else {}
        students_map = {s.id: s for s in students_list}
    else:
        students_map = {}
        classes_map = {}

    # Build preview rows
    rows = []
    for sc in all_scores:
        student = students_map.get(sc.student_id)
        cls = classes_map.get(student.class_id) if student else None
        rows.append({
            "student_id": student.student_id if student else "",
            "student_name": student.name if student else "",
            "gender": student.gender.value if student else "",
            "class": f"{cls.grade}{cls.name}" if cls else "",
            "event_name": event_map[sc.event_id].name if sc.event_id in event_map else "",
            "raw_value": sc.raw_value,
            "earned_score": sc.earned_score,
            "test_date": sc.test_date.isoformat()
        })

    return {"rows": rows, "total": len(rows)}

@router.get("/export/download")
def export_download(
    scope: str = Query(...),
    class_id: Optional[int] = Query(None),
    student_id: Optional[int] = Query(None),
    event_ids: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    mode: str = Query("all"),
    format: str = Query("xlsx"),
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin_flexible)
):
    """Download export file (xlsx or txt)."""
    sid = _get_admin_school(current)
    try:
        return _do_export_download(scope, class_id, student_id, event_ids, date_from, date_to, mode, format, sid, db)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"{type(e).__name__}: {e}")

def _do_export_download(scope, class_id, student_id, event_ids, date_from, date_to, mode, format, sid, db):
    # Reuse preview logic
    event_id_list = [int(x) for x in event_ids.split(",")] if event_ids else None
    q = _maybe_filter(db.query(Score), Score, sid)
    if scope == "class" and class_id:
        students_in_class = db.query(Student).filter(Student.class_id == class_id).all()
        q = q.filter(Score.student_id.in_([s.id for s in students_in_class]))
    elif scope == "student" and student_id:
        q = q.filter(Score.student_id == student_id)
    if event_id_list:
        q = q.filter(Score.event_id.in_(event_id_list))
    if date_from:
        q = q.filter(Score.test_date >= date.fromisoformat(date_from))
    if date_to:
        q = q.filter(Score.test_date <= date.fromisoformat(date_to))

    all_scores = q.order_by(Score.test_date.desc(), Score.student_id).all()
    events = db.query(SportEvent).filter(SportEvent.school_id == sid).all()
    event_map = {e.id: e for e in events}

    if mode == "best":
        best = {}
        for sc in all_scores:
            key = (sc.student_id, sc.event_id)
            if key not in best or sc.earned_score > best[key].earned_score:
                best[key] = sc
        all_scores = sorted(best.values(), key=lambda x: (x.student_id, x.event_id))
    elif mode == "latest":
        latest = {}
        for sc in all_scores:
            key = (sc.student_id, sc.event_id)
            if key not in latest:
                latest[key] = sc
        all_scores = sorted(latest.values(), key=lambda x: (x.student_id, x.event_id))

    # Preload students and classes
    student_ids = {sc.student_id for sc in all_scores} if all_scores else set()
    if student_ids:
        students_list = db.query(Student).filter(Student.id.in_(student_ids)).all()
        class_ids = {s.class_id for s in students_list}
        classes_map = {c.id: c for c in db.query(Class).filter(Class.id.in_(class_ids)).all()} if class_ids else {}
        students_map = {s.id: s for s in students_list}
    else:
        students_map = {}
        classes_map = {}

    scope_label = {"school": "全校", "class": "班级", "student": "个人"}.get(scope, "")
    mode_label = {"all": "全部", "best": "最优", "latest": "最近"}.get(mode, "")

    if format == "txt":
        lines = []
        for sc in all_scores:
            student = students_map.get(sc.student_id)
            cls = classes_map.get(student.class_id) if student else None
            name = student.name if student else ""
            evt = event_map[sc.event_id].name if sc.event_id in event_map else ""
            lines.append(f"{name}\t{evt}\t{sc.raw_value}\t{sc.earned_score}分\t{sc.test_date}")
        content = "\n".join(lines)
        return StreamingResponse(
            iter([content]), media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename=export_{scope}_{mode}.txt"}
        )

    # xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{scope_label}成绩{mode_label}"
    ws.append(["学号", "姓名", "性别", "班级", "项目", "成绩", "得分", "测试日期"])
    for sc in all_scores:
        student = students_map.get(sc.student_id)
        cls = classes_map.get(student.class_id) if student else None
        ws.append([
            student.student_id if student else "", student.name if student else "",
            student.gender.value if student else "", f"{cls.grade}{cls.name}" if cls else "",
            event_map[sc.event_id].name if sc.event_id in event_map else "",
            sc.raw_value, sc.earned_score, sc.test_date.isoformat()
        ])
    buffer = BytesIO()
    wb.save(buffer); buffer.seek(0)
    return StreamingResponse(
        buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=export_{scope}_{mode}.xlsx"}
    )


@router.get("/student-list/{class_id}")
def get_class_students(
    class_id: int,
    event_id: int = Query(None),
    db: Session = Depends(get_db),
    current: Admin = Depends(get_current_admin)
):
    sid = _get_admin_school(current)
    cls = db.query(Class).filter(Class.id == class_id, Class.school_id == sid).first()
    if not cls:
        raise HTTPException(404, "班级不存在")
    q = db.query(Student).filter(Student.class_id == class_id)
    if event_id:
        event = db.query(SportEvent).filter(
            SportEvent.id == event_id, SportEvent.school_id == sid
        ).first()
        if event and event.gender.value != "both":
            q = q.filter(Student.gender == event.gender)
    students = q.order_by(Student.student_id).all()
    return [{"id": s.id, "student_id": s.student_id, "name": s.name, "gender": s.gender.value} for s in students]

@router.get("/backup-all")
def backup_all_data(db: Session = Depends(get_db), current: Admin = Depends(get_school_admin)):
    """Backup current school data as Excel (7 sheets)."""
    sid = _get_admin_school(current)
    wb = openpyxl.Workbook()

    # Sheet 1: Classes
    ws1 = wb.active
    ws1.title = "班级信息"
    ws1.append(["ID", "年级", "班级名", "学校ID"])
    classes = db.query(Class).filter(Class.school_id == sid).order_by(Class.grade, Class.name).all()
    for c in classes:
        ws1.append([c.id, c.grade, c.name, c.school_id])

    # Sheet 2: Sport Events
    ws2 = wb.create_sheet("体育项目")
    ws2.append(["ID", "名称", "性别", "越大越好", "单位", "输入格式", "排序", "学校ID"])
    events = db.query(SportEvent).filter(SportEvent.school_id == sid).order_by(SportEvent.sort_order).all()
    for e in events:
        ws2.append([e.id, e.name, e.gender.value if e.gender else "both", e.higher_better, e.unit, e.input_format.value if e.input_format else "decimal_seconds", e.sort_order, e.school_id])

    # Sheet 3: Scoring Standards
    ws3 = wb.create_sheet("评分标准")
    ws3.append(["ID", "项目ID", "性别", "分数", "标准值"])
    event_ids = [e.id for e in events]
    standards = db.query(ScoringStandard).filter(ScoringStandard.event_id.in_(event_ids)).order_by(ScoringStandard.event_id, ScoringStandard.score.desc()).all() if event_ids else []
    for std in standards:
        ws3.append([std.id, std.event_id, std.gender.value if std.gender else "both", std.score, std.standard_value])

    # Sheet 4: Students
    ws4 = wb.create_sheet("学生信息")
    ws4.append(["ID", "学号", "姓名", "性别", "班级ID", "密码哈希"])
    class_ids = [c.id for c in classes]
    students = db.query(Student).filter(Student.class_id.in_(class_ids)).order_by(Student.student_id).all() if class_ids else []
    for s in students:
        ws4.append([s.id, s.student_id, s.name, s.gender.value if s.gender else "M", s.class_id, s.password_hash])

    # Sheet 5: Scores
    ws5 = wb.create_sheet("成绩记录")
    ws5.append(["ID", "学生ID", "项目ID", "成绩", "得分", "测试日期", "录入人ID", "学校ID"])
    scores = db.query(Score).filter(Score.school_id == sid).order_by(Score.test_date.desc()).all()
    for sc in scores:
        ws5.append([sc.id, sc.student_id, sc.event_id, sc.raw_value, sc.earned_score, sc.test_date.isoformat(), sc.recorder_id or "", sc.school_id])

    # Sheet 6: Admins
    ws6 = wb.create_sheet("管理员")
    ws6.append(["ID", "用户名", "密码哈希", "超管", "显示名", "学校ID"])
    admins = db.query(Admin).filter((Admin.school_id == sid) | (Admin.is_super == True)).order_by(Admin.id).all()
    for a in admins:
        ws6.append([a.id, a.username, a.password_hash, a.is_super, a.display_name, a.school_id or ""])

    # Sheet 7: System Config
    ws7 = wb.create_sheet("系统设置")
    ws7.append(["键", "值", "学校ID"])
    configs = db.query(SystemConfig).filter(SystemConfig.school_id == sid).all()
    for cfg in configs:
        ws7.append([cfg.key, cfg.value, cfg.school_id])

    buffer = BytesIO(); wb.save(buffer); buffer.seek(0)
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=sports_backup.xlsx"})

@router.post("/restore-all")
def restore_all_data(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current: Admin = Depends(get_school_admin)
):
    """Restore current school data from backup Excel file."""
    sid = _get_admin_school(current)
    try:
        wb = openpyxl.load_workbook(file.file, data_only=True)
    except Exception:
        raise HTTPException(400, "无法读取文件，请确认是备份Excel文件")

    # Read sheets by index (order is fixed regardless of encoding)
    def get_sheet_data(idx):
        if idx < len(wb.sheetnames):
            return list(wb[wb.sheetnames[idx]].iter_rows(min_row=2, values_only=True))
        return []

    # Sheet order: 0=班级信息, 1=体育项目, 2=评分标准, 3=学生信息, 4=成绩记录, 5=管理员, 6=系统设置
    classes_data = [{"id": int(r[0]), "grade": str(r[1]), "name": str(r[2]), "school_id": int(r[3]) if len(r) > 3 and r[3] else sid} for r in get_sheet_data(0) if r[0] is not None]
    events_data = []
    for r in get_sheet_data(1):
        if r[0] is not None:
            events_data.append({
                "id": int(r[0]), "name": str(r[1]), "gender": Gender(str(r[2])) if r[2] else Gender.both,
                "higher_better": r[3] in (True, "True", "true", 1, "1"),
                "unit": str(r[4]), "input_format": InputFormat(str(r[5])) if r[5] else InputFormat.decimal_seconds,
                "sort_order": int(r[6] or 0), "school_id": int(r[7]) if len(r) > 7 and r[7] else sid
            })
    standards_data = []
    for r in get_sheet_data(2):
        if r[0] is not None:
            standards_data.append({
                "id": int(r[0]), "event_id": int(r[1]),
                "gender": Gender(str(r[2])) if r[2] else Gender.both,
                "score": int(r[3]), "standard_value": str(r[4])
            })
    students_data = []
    for r in get_sheet_data(3):
        if r[0] is not None:
            students_data.append({
                "id": int(r[0]), "student_id": str(r[1]), "name": str(r[2]),
                "gender": Gender(str(r[3])) if r[3] else Gender.M, "class_id": int(r[4]),
                "password_hash": str(r[5]) if len(r) > 5 and r[5] else ""
            })
    scores_data = []
    for r in get_sheet_data(4):
        if r[0] is not None:
            scores_data.append({
                "id": int(r[0]), "student_id": int(r[1]), "event_id": int(r[2]),
                "raw_value": str(r[3]), "earned_score": int(r[4]),
                "test_date": date.fromisoformat(str(r[5])) if r[5] else date.today(),
                "recorder_id": int(r[6]) if r[6] and str(r[6]).strip() else None,
                "school_id": int(r[7]) if len(r) > 7 and r[7] else sid
            })
    admins_data = []
    for r in get_sheet_data(5):
        if r[0] is not None:
            admins_data.append({
                "id": int(r[0]), "username": str(r[1]), "password_hash": str(r[2]),
                "is_super": r[3] in (True, "True", "true", 1, "1"), "display_name": str(r[4]),
                "school_id": int(r[5]) if len(r) > 5 and r[5] else None
            })
    configs_data = []
    for r in get_sheet_data(6):
        if r[0] is not None:
            configs_data.append({
                "key": str(r[0]), "value": str(r[1]),
                "school_id": int(r[2]) if len(r) > 2 and r[2] else sid
            })

    try:
        # Delete current school's data in reverse dependency order
        db.query(Score).filter(Score.school_id == sid).delete(synchronize_session=False)
        db.query(ScoringStandard).filter(ScoringStandard.event_id.in_(
            db.query(SportEvent.id).filter(SportEvent.school_id == sid)
        )).delete(synchronize_session=False)
        db.query(Student).filter(Student.class_id.in_(
            db.query(Class.id).filter(Class.school_id == sid)
        )).delete(synchronize_session=False)
        db.query(SportEvent).filter(SportEvent.school_id == sid).delete(synchronize_session=False)
        db.query(Class).filter(Class.school_id == sid).delete(synchronize_session=False)
        db.query(Admin).filter(Admin.school_id == sid).delete(synchronize_session=False)
        db.query(SystemConfig).filter(SystemConfig.school_id == sid).delete(synchronize_session=False)
        db.flush()

        # Insert in dependency order
        for d in classes_data:
            db.add(Class(**d))
        for d in events_data:
            db.add(SportEvent(**d))
        for d in standards_data:
            db.add(ScoringStandard(**d))
        for d in students_data:
            db.add(Student(**d))
        for d in scores_data:
            db.add(Score(**d))
        for d in admins_data:
            db.add(Admin(**d))
        for d in configs_data:
            db.add(SystemConfig(**d))
        db.flush()

        # Reset auto-increment sequences (PostgreSQL only; SQLite handles it automatically)
        if 'postgresql' in settings.database_url:
            for tbl in ['classes', 'sport_events', 'scoring_standards', 'students', 'scores', 'admins']:
                db.execute(text(f"SELECT setval('{tbl}_id_seq', COALESCE((SELECT MAX(id) FROM {tbl}), 1))"))

        db.commit()
        return {"ok": True, "classes": len(classes_data), "events": len(events_data), "standards": len(standards_data),
                "students": len(students_data), "scores": len(scores_data), "admins": len(admins_data), "configs": len(configs_data)}
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"恢复失败: {e}")

@router.post("/clear-all")
def clear_all_scores(
    data: ClearAllRequest,
    db: Session = Depends(get_db),
    current: Admin = Depends(get_super_admin)
):
    """Delete all score records for current school (super admin only, password required)."""
    if not verify_password(data.password, current.password_hash):
        raise HTTPException(403, "密码错误")
    sid = _get_admin_school(current)
    count = _maybe_filter(db.query(Score), Score, sid).count()
    _maybe_filter(db.query(Score), Score, sid).delete(synchronize_session=False)
    db.commit()
    return {"ok": True, "deleted": count}
